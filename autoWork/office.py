# coding=utf-8
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font
from openpyxl.styles import colors
from openpyxl.styles import Font
import os
from python.Image import imgTest

#http://nullege.com/codes/search/openpyxl.style.Color
# 打开文档
fathPath = "jinlong/Desktop/myfile/myexcel.xlsx"
def testExcel():
    path = "/Users/cl/Desktop/icrawlerLearn/myfile/myexcel.xlsx"

    # imgPath = father_path + "/Image"+"/dog.png"
    # print("imgPath",imgPath)
    (colorArr, imW, imH) = imgTest.getAllColorsFromImage()
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        print(wb.get_sheet_names())
        #获取当前工作表
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

        print(sheet.title)
        worksheet = wb.get_active_sheet()
        print(sheet['A2'].value)
        c = sheet["A1"]
        print(c.value)
        a = sheet.cell(row=1,column=1).value
        print(a)
        c = sheet.cell(row=3,column=1).value
        print(type(c)) #<class 'int'>
        for var in sheet.iter_rows():
            print(var[0].value)
        countNumber = 0
        for value in sheet.get_cell_collection():
            if value.value != None:
                countNumber +=1
                print(value.value,(value.row,value.column))

        print("统计数据为:",countNumber)

        wb = Workbook()
        ws = wb.active
        # 创建一个worksheet在最后面的位置(默认创建的工作表会在最后面，除非你指定位置)
        ws1 = wb.create_sheet("mysheet1")
        # 创建一个工作表在最前面的位置，注意编号
        ws2 = wb.create_sheet("mysheet", 0)
        ws.title = "New Title"
        ws.sheet_properties.tabColor = "1072BA"
        # 当然，如果你更改了工作表（worksheet）的名字，你可以用切换过去
        # ws = wb["New Title"]
        #用openpyxl.workbook.Workbook.copy_worksheet()来复制工作表
        source = wb.active
        target = wb.copy_worksheet(source)
        ws["A4"] = 4
        colorIndex = 0
        d = ws.cell(row=4, column=2, value=10)
        for i in range(1, imW):
            for j in range(1, imH):
                cell = ws.cell(row=i, column=j,value="%d*%d=%d"%(i,j,i*j))
                ft = Font(color=colors.RED)
                (r,g,b,a) = colorArr[colorIndex]
                colorHexStr = "%02x%02x%02x" % (r, g, b)

                progerss = colorIndex/(imH*imW)
                strp = "##"
                print(strp *int((10*progerss)),"%02d"%(100*progerss),"%")

                ft = Font(color=colorHexStr)
                cell.font = ft
                colorIndex +=1
                # cell.font.italic = False #is not allowed 对象在指定Styles后是不允许被更改的
        # 可用切片来访问一个单元格范围
        cell_range = ws["A1":"C2"]
        colc = ws["C"]
        col_range = ws["C:D"]
        row10 = ws[10]
        row_range = ws[5:10]
        savePath = fathPath+"/ws.xlsx"
        print(savePath)
        wb.save(savePath)
        print("end save Excel")
testExcel()
def testWorld(path,text):
    if os.path.exists(path):
        print("exist path")

    else:
        return "path is error"

